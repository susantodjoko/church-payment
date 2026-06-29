import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from collections import defaultdict

_THIN = Side(style='thin')
_TOP_BORDER = Border(top=_THIN)
_IDR = '#,##0'

_LK_MONTH_NAMES = ['JAN', 'FEB', 'MRT', 'APR', 'MEI', 'JUN',
                   'JUL', 'AGT', 'SEP', 'OKT', 'NOV', 'DES']
_LK_MONTH_FULL = ['JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI',
                  'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER']
_LK_COL_HEADERS = [
    'Bulan', '', 'Tgl', 'NO BUKTI', 'Rincian Transaksi', 'Kode',
    'Lingkungan Penyetor/Penerima', 'Debit', 'Kredit', 'Saldo', 'Keterangan',
]
_LK_COL_WIDTHS = [6, 4, 12, 10, 26, 6, 26, 14, 14, 14, 12]


def _header_style(ws, row, headers):
    fill = PatternFill(fill_type='solid', fgColor='1F3864')
    font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def build_monthly_excel(payments, month, year):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Laporan {month}-{year}'
    ws.cell(1, 1, f'Laporan Pembayaran Bulan {month}/{year}').font = Font(bold=True, size=14)
    headers = ['No', 'Anggota', 'Lingkungan', 'Wilayah', 'Jenis', 'Jumlah (Rp)', 'Tgl Terima', 'Dicatat oleh']
    _header_style(ws, 3, headers)
    for i, p in enumerate(payments, 1):
        if p.member:
            subject = p.member.full_name
            lingkungan = p.member.lingkungan.name
            wilayah = p.member.lingkungan.wilayah.name
        else:
            subject = str(p.keluarga)
            lingkungan = p.keluarga.lingkungan.name if p.keluarga else '-'
            wilayah = p.keluarga.lingkungan.wilayah.name if p.keluarga else '-'
        ws.append([
            i,
            subject,
            lingkungan,
            wilayah,
            p.payment_type.name,
            float(p.amount),
            p.date_received.strftime('%d/%m/%Y %H:%M'),
            p.recorded_by.get_full_name() or p.recorded_by.username,
        ])
    ws.append([])
    ws.append(['', '', '', '', 'TOTAL', sum(float(p.amount) for p in payments)])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_annual_excel(year, monthly_totals, member_summary):
    wb = openpyxl.Workbook()

    # Sheet 1: month-by-month totals
    ws1 = wb.active
    ws1.title = 'Rekapitulasi Bulanan'
    ws1.cell(1, 1, f'Laporan Tahunan {year}').font = Font(bold=True, size=14)
    _header_style(ws1, 3, ['Bulan', 'Total Pembayaran (Rp)', 'Jumlah Transaksi'])
    for row in monthly_totals:
        ws1.append(row)

    # Sheet 2: per-member summary
    ws2 = wb.create_sheet('Per Anggota')
    _header_style(ws2, 1, ['Anggota', 'Lingkungan', 'Bulan Dibayar', 'Total (Rp)', 'Belum Bayar (bulan)'])
    for row in member_summary:
        ws2.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_unpaid_excel(unpaid_members, month, year, payment_type_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Belum Bayar'
    ws.cell(1, 1, f'Anggota Belum Bayar — {payment_type_name} {month}/{year}').font = Font(bold=True, size=14)
    _header_style(ws, 3, ['No', 'ID Anggota', 'Nama', 'Lingkungan', 'Wilayah', 'Telepon'])
    for i, m in enumerate(unpaid_members, 1):
        ws.append([
            i, m.member_id, m.full_name,
            m.lingkungan.name, m.lingkungan.wilayah.name,
            m.phone or '-',
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── LK PKSS helpers ──────────────────────────────────────────────────────────

def _lk_col_headers(ws, row):
    fill = PatternFill(fill_type='solid', fgColor='1F3864')
    font = Font(color='FFFFFF', bold=True, size=9)
    for col, h in enumerate(_LK_COL_HEADERS, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def _lk_total(ws, row, label, debit, kredit, saldo):
    ws.cell(row, 5, label).font = Font(bold=True)
    for col, val in [(8, debit), (9, kredit), (10, saldo)]:
        c = ws.cell(row, col, val)
        c.font = Font(bold=True)
        c.number_format = _IDR
    for col in range(1, 12):
        ws.cell(row, col).border = _TOP_BORDER


def _lk_monthly_sheet(wb, month_idx, year, entries, lng_code_map, paroki_name):
    """Build one monthly KAS+BANK sheet. entries = list of (Payment, Lingkungan)."""
    ws = wb.create_sheet(_LK_MONTH_NAMES[month_idx])
    for i, w in enumerate(_LK_COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # Header block
    ws.cell(1, 2, f'Paroki : {paroki_name}').font = Font(bold=True)
    ws.cell(2, 2, 'LAPORAN KEUANGAN PKSS').font = Font(bold=True)
    ws.cell(3, 2, f'BULAN : {_LK_MONTH_FULL[month_idx]} {year}').font = Font(bold=True)

    # KAS column headers & opening balance
    _lk_col_headers(ws, 5)
    ws.cell(6, 3, 'KAS').font = Font(bold=True)
    ws.cell(6, 10, 0).number_format = _IDR

    # KAS entries from app data
    cur = 7
    saldo = 0.0
    for n, (p, lng) in enumerate(entries, 1):
        amount = float(p.amount)
        saldo += amount
        ws.cell(cur, 2, n)
        ws.cell(cur, 3, p.date_received.strftime('%d/%m/%Y'))
        ws.cell(cur, 5, 'IURAN PKSS')
        ws.cell(cur, 6, lng_code_map.get(lng.id, ''))
        ws.cell(cur, 7, lng.name)
        ws.cell(cur, 8, amount).number_format = _IDR
        ws.cell(cur, 10, saldo).number_format = _IDR
        cur += 1

    # Pad KAS section to at least row 29 (matches template row layout)
    while cur <= 29:
        ws.cell(cur, 10, saldo).number_format = _IDR
        cur += 1

    # TOTAL KAS
    total_kas_row = cur
    total_debit = sum(float(p.amount) for p, _ in entries)
    _lk_total(ws, total_kas_row, 'TOTAL KAS', total_debit, 0, saldo)
    cur += 2  # leave one blank row

    # BANK section header
    ws.cell(cur, 3, f'BANK BCA AN. BGKP {paroki_name} (PASTORAN)').font = Font(bold=True)
    cur += 1
    _lk_col_headers(ws, cur)
    cur += 1
    ws.cell(cur, 10, 0).number_format = _IDR  # opening balance
    cur += 1

    # 17 empty numbered rows for manual bank entry
    for i in range(1, 18):
        ws.cell(cur, 2, i)
        cur += 1

    # TOTAL BANK
    bank_total_row = cur
    _lk_total(ws, bank_total_row, 'TOTAL BANK', 0, 0, 0)
    cur += 2

    # TOTAL KAS & BANK
    _lk_total(ws, cur, 'TOTAL KAS & BANK', total_debit, 0, saldo)
    ws.cell(cur, 5).font = Font(bold=True, size=11)


def build_lk_pkss_excel(year, lingkungan_list, pkss_payments, paroki_name='ST. STEFANUS'):
    """
    Generate LK PKSS Excel matching the Laporan Keuangan PKSS template.

    Monthly sheets (JAN-DES): KAS section auto-filled from pkss_payments;
    BANK section left as empty template for manual entry.
    REKAP sheet: all lingkungan × 12 months, auto-totalled.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Group payments: monthly_map[month] = [(Payment, Lingkungan)]
    # total_map[(lng_id, month)] = float total
    monthly_map = defaultdict(list)
    total_map = defaultdict(float)

    for p in pkss_payments:
        lng = p.member.lingkungan if p.member else (
            p.keluarga.lingkungan if p.keluarga else None
        )
        if lng is None:
            continue
        monthly_map[p.period_month].append((p, lng))
        total_map[(lng.id, p.period_month)] += float(p.amount)

    for m in range(1, 13):
        monthly_map[m].sort(key=lambda x: x[0].date_received)

    # Sequential kode (1-based) per the lingkungan ordering
    lng_code_map = {lng.id: i + 1 for i, lng in enumerate(lingkungan_list)}

    # Build monthly sheets
    for month_idx in range(12):
        _lk_monthly_sheet(
            wb, month_idx, year,
            monthly_map[month_idx + 1],
            lng_code_map, paroki_name,
        )

    # Build REKAP sheet
    ws_r = wb.create_sheet('REKAP')
    ws_r.column_dimensions['A'].width = 6
    ws_r.column_dimensions['B'].width = 24
    ws_r.column_dimensions['C'].width = 24
    for col_i in range(4, 17):
        ws_r.column_dimensions[ws_r.cell(1, col_i).column_letter].width = 13
    ws_r.column_dimensions[ws_r.cell(1, 17).column_letter].width = 14

    ws_r.cell(1, 2, f'PENERIMAAN IURAN PKSS {year}').font = Font(bold=True, size=14)

    # REKAP header
    rekap_hdr = ['KODE', 'WIL/STASI', 'LINGKUNGAN'] + _LK_MONTH_NAMES + ['TOTAL']
    fill = PatternFill(fill_type='solid', fgColor='1F3864')
    font_hdr = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(rekap_hdr, 1):
        c = ws_r.cell(3, col, h)
        c.fill = fill
        c.font = font_hdr
        c.alignment = Alignment(horizontal='center')

    # Data rows
    data_row = 4
    for i, lng in enumerate(lingkungan_list):
        kode = i + 1
        monthly_totals = [total_map.get((lng.id, m), 0.0) for m in range(1, 13)]
        row_vals = [kode, lng.wilayah.name, lng.name] + monthly_totals + [sum(monthly_totals)]
        for col, val in enumerate(row_vals, 1):
            c = ws_r.cell(data_row, col, val)
            if col >= 4:
                c.number_format = _IDR
                c.alignment = Alignment(horizontal='right')
        data_row += 1

    # T O T A L row
    grand_by_month = [
        sum(total_map.get((lng.id, m), 0.0) for lng in lingkungan_list)
        for m in range(1, 13)
    ]
    grand_total = sum(grand_by_month)
    total_vals = ['', '', 'T O T A L'] + grand_by_month + [grand_total]
    for col, val in enumerate(total_vals, 1):
        c = ws_r.cell(data_row, col, val)
        c.font = Font(bold=True)
        if col >= 4:
            c.number_format = _IDR
        c.border = _TOP_BORDER
    data_row += 1

    # Pendaftaran anggota baru (empty — manual entry)
    ws_r.cell(data_row, 2, 'Pendaftaran anggota baru').font = Font(italic=True)
    for col in range(4, 17):
        ws_r.cell(data_row, col, 0).number_format = _IDR
    data_row += 1

    # Total Penerimaan (PKKS only; user adds registration fees manually)
    ws_r.cell(data_row, 3, 'Total Penerimaan').font = Font(bold=True)
    for m_i, val in enumerate(grand_by_month):
        c = ws_r.cell(data_row, 4 + m_i, val)
        c.font = Font(bold=True)
        c.number_format = _IDR
    c = ws_r.cell(data_row, 16, grand_total)
    c.font = Font(bold=True)
    c.number_format = _IDR

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
